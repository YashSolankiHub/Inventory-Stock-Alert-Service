from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.utils.validators import PwdContext
from fastapi import HTTPException, status, UploadFile
from sqlalchemy import and_, or_, func
from app.services.common_service import CommonService
from openpyxl.styles import Alignment
import os

from app.models.users import User as UserModel
from app.models.categories import Category as CategoryModel

from app.schemas.user import UserResponseSchema
from app.schemas.general_response import ErrorResponse, StandardResponse, ErrorDetail
from datetime import datetime, timezone
from app.utils.logging import LoggingService
from typing import TYPE_CHECKING
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.exc import DataError
from app.exceptions.database import *
from app.exceptions.auth import *
from app.exceptions.common import *
from app.models.products import Product as ProductModel
from app.utils.helper import generate_sku
from app.models.warehouses import Warehouse as WarehouseModel
from app.enums.enums import UserRoles
from app.models.bins import Bin as BinModel
from app.models.inventory_item import InventoryItem as InventoryItemModel
from app.models.purchase_order_items import POItem as POItemModel
from app.exceptions.bin import *
from app.exceptions.purchase_order import *
from app.models.received_po_item import ReceivedPOItem as ReceivedPOItemModel
from app.models.purchase_orders import PurchaseOrder as PurchaseOrderModel
from app.enums.enums import PurchaseOrderStatus
from openpyxl import Workbook
from io import BytesIO
from fastapi.responses import StreamingResponse



import uuid

logger = LoggingService(__name__).get_logger() 


class InventoryItemService(CommonService):
    def __init__(self, db:Session):
        self.db = db
        CommonService.__init__(self,db, InventoryItemModel)

    def add_item_in_inventory(self,py_model:BaseModel):
        pydantic_data = py_model.model_dump()
        logger.info(f"Adding item in inventory with data: {pydantic_data}")

        purchase_order_record = self.db.get(PurchaseOrderModel, pydantic_data['po_id'])

        #raise exception if purchase_order_record not found
        if not purchase_order_record:
            logger.warning(f"Purchase Order with id {pydantic_data['po_id']} not exists")
            raise NotFoundException(f"Purchase Order with id {pydantic_data['po_id']} not exists")
        
        if purchase_order_record.status != PurchaseOrderStatus.RECEIVED:
            logger.warning(f"Purchase Order with id {pydantic_data['po_id']} status is not received")
            raise InvalidStatusException(f"Cannot proceed. Purchase Order must be in 'RECEIVED' state, currently in {purchase_order_record.status.value}")
        

        received_item_record = self.db.query(ReceivedPOItemModel).filter_by(product_id = pydantic_data['product_id']).first()
        
        #raise exception if received_item_record not found
        if not received_item_record:
            logger.warning(f"Received item with product id{pydantic_data['product_id']} not found!")
            raise NotFoundException(f"Received item with product id {pydantic_data['product_id']} not found!")
        
        if pydantic_data['qty'] > received_item_record.qty:
            logger.warning(f"Product with id {pydantic_data['product_id']} does not receive {pydantic_data['qty']}. Actual unit is : {received_item_record.qty}")
            raise BadRequestException(msg = f"Product with id {pydantic_data['product_id']} does not receive {pydantic_data['qty']} units. Actual unit is : {received_item_record.qty}", code = "EXPECTED_QUANTITY_EXCEEDS")
        
        bin_record = self.db.query(BinModel).filter_by(id = pydantic_data['bin_id'], warehouse_id = purchase_order_record.warehouse_id).first()

        #raise exception if bin record not found
        if not bin_record:
            logger.warning(f"Bin with id{pydantic_data['bin_id']} not found!")
            raise NotFoundException(f"Bin with id{pydantic_data['bin_id']} not found!")
        
        #raise exception if request qty is greater than available capicity
        if pydantic_data['qty'] > bin_record.available_units:
            logger.warning(f"The quantity ({pydantic_data['qty']}) exceeds the bin's available capacity ({bin_record.max_units})")
            raise BinCapacityExceededException(bin_record.max_units, pydantic_data['qty'])
        
        #add qty to bin current stock
        bin_record.current_stock_units = bin_record.current_stock_units + pydantic_data['qty']
        #substract qty from available unit
        bin_record.available_units = bin_record.available_units - pydantic_data['qty']

        pydantic_data["product_id"] = received_item_record.product_id
        pydantic_data["sku"] = received_item_record.sku
        pydantic_data["warehouse_id"] = purchase_order_record.warehouse_id
        logger.info(f"Adding item in inventory with data: {pydantic_data}")
        pydantic_data.pop("po_id")
        logger.info(f"Deleting po_id from pydantic data: {pydantic_data}")
        inventory_item = self.create_record(pydantic_data)
        logger.info(f"Deleting received item with id {received_item_record.id}")

        #substract received item qty which is added into bin
        received_item_record.qty = received_item_record.qty - pydantic_data['qty']

        try:
            if received_item_record.qty == 0:
                self.db.delete(received_item_record)
            self.db.commit()
        except SQLAlchemyError as e:
            self.db.rollback()
            raise DataBaseError(e)
        return inventory_item
    

    def transfer_stock_between_warehouse(self, py_model:BaseModel):
        logger.info("Start process for transfering stock")
        pydantic_data = py_model.model_dump()
        logger.info(f"Received data for transfering : {pydantic_data}")

        warehouse_record = self.db.get(WarehouseModel, pydantic_data['from_warehouse_id'])

        #raise exception if warehouse not found
        if not warehouse_record:
            logger.warning(f"warehouse with id {pydantic_data['from_warehouse_id']} not exists")
            raise FileNotFoundError(f"warehouse with id {pydantic_data['from_warehouse_id']} not exists")
        
        product_record = self.db.get(InventoryItemModel, pydantic_data['product_id'])

        if not product_record:
            logger.warning(f"Product with id {pydantic_data['product_id']}not exists")
            raise FileNotFoundError(f"Product with id {pydantic_data['product_id']}not exists")
        


    def export_inventory_report(self, warehouse_id=None):
        logger.info("Start process for exporting inventory report")

        product_records = self.db.query(InventoryItemModel)

        if warehouse_id:
            warehouse_record = self.db.query(WarehouseModel).filter_by(id = warehouse_id).first()
            #raise exception if warehouse not exists
            if not warehouse_record :
                logger.warning(f"Warehouse with id  {warehouse_id} not exists!")
                raise NotFoundException(f"Warehouse with id  {warehouse_id} not exists!")
            
            all_product_records = product_records.filter_by(warehouse_id=warehouse_id).all()
            low_qty_product_records = (
                product_records
                .join(ProductModel)
                .filter(
                    InventoryItemModel.warehouse_id == warehouse_id,
                    InventoryItemModel.qty < ProductModel.threshold_qty
                )
                .all()
            )

            if not all_product_records:
                logger.warning(f"Warehouse with id {warehouse_id} does not have any product stocks")
                raise NotFoundException(f"Warehouse with id {warehouse_id} does not have any product stocks")
        else:
            all_product_records = product_records.all()
            low_qty_product_records = (
                product_records
                .join(ProductModel)
                .filter(InventoryItemModel.qty < ProductModel.threshold_qty)
                .all()
            )

            if not all_product_records:
                logger.warning("No inventory records found")
                raise NotFoundException("No inventory records found")

        # excel worbook created
        wb = Workbook()
        current_ws = wb.active
        current_ws.title = "Current Inventory"

        #write headers
        current_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        current_ws.cell(row=1, column=1).value = f"Warehouse ID: {warehouse_id or 'All Warehouses'}"
        current_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        current_ws.append(["Product ID", "SKU", "Warehouse ID", "Bin ID", "Qty"])

        for item in all_product_records:
            current_ws.append([
                str(item.product_id),
                item.sku,
                str(item.warehouse_id),
                str(item.bin_id),
                item.qty,
            ])

        if low_qty_product_records:
            low_stock_ws = wb.create_sheet("Low Stock Items")
            low_stock_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            low_stock_ws.cell(row=1, column=1).value = f"Warehouse ID: {warehouse_id or 'All Warehouses'}"
            low_stock_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            low_stock_ws.append(["Product ID", "SKU", "Warehouse ID", "Bin ID", "Qty"])

            for item in low_qty_product_records:
                low_stock_ws.append([
                    str(item.product_id),
                    item.sku,
                    str(item.warehouse_id),
                    str(item.bin_id),
                    item.qty,
                ])


        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)

        filename = f"{uuid.uuid4()}_inventory_report_{warehouse_id or 'all'}.xlsx"
        file_path = os.path.join(reports_dir, filename)

        wb.save(file_path)

        return file_path

    
    














































































    # def export_inventory_report(self, warehouse_id):
    #     logger.info(f"start process for exportin reports")

    #     product_records = self.db.query(InventoryItemModel)

    #     if warehouse_id:
    #         all_product_records = product_records.filter_by(warehouse_id = warehouse_id).all()

    #         low_qty_product_records = product_records.join(ProductModel).filter(InventoryItemModel.qty < ProductModel.threshold_qty).all()

    #         if not all_product_records:
    #             logger.warning(f"Warhouse with id {warehouse_id} does not have any product stocks")
    #             raise NotFoundException(f"Warhouse with id {warehouse_id} does not have any product stocks")
            
    #         wb = Workbook()
    #         current_ws = wb.active
    #         current_ws.title = "Current Inventory"

    #         if low_qty_product_records:
    #             low_stock_ws = wb.create_sheet("Low Stock Items")
    #             low_stock_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    #             low_stock_ws.cell(row=1, column=1).value = f"Warehouse ID: {warehouse_id or 'All Warehouses'}"
    #             low_stock_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    #             low_stock_ws.append(["Product ID", "SKU", "Warehouse ID", "Bin ID", "Qty"])

    #             for item in low_qty_product_records:
    #                 low_stock_ws.append([
    #                     str(item.product_id),
    #                     item.sku,
    #                     str(item.warehouse_id),
    #                     str(item.bin_id),
    #                     item.qty,
    #                 ])

    #         current_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    #         current_ws.cell(row=1, column=1).value = f"Warehouse ID: {warehouse_id}"
    #         current_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    #         current_ws.append(["Product ID", "SKU", "Warehouse ID", "Bin ID", "Qty"])
    #         for item in all_product_records:
    #             current_ws.append([
    #                 str(item.product_id),
    #                 item.sku,
    #                 str(item.warehouse_id),
    #                 str(item.bin_id),
    #                 item.qty,
    #             ])
            
    #         # Write to memory
    #         stream = BytesIO()
    #         wb.save(stream)
    #         stream.seek(0)

    #         reports_dir = "reports"
    #         os.makedirs(reports_dir, exist_ok=True)

    #         file_path = os.path.join(reports_dir, f"{uuid.uuid4()}_inventory_report_{warehouse_id}.xlsx")
    #         wb.save(file_path)

    #     return file_path






        

        



    

