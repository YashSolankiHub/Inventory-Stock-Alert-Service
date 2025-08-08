from app.core.celery_app import celery_app
from datetime import datetime
from app.db.database import db
from sqlalchemy.orm import Session
from app.models.inventory_item import InventoryItem
from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
import os
from app.utils.email import Email
import uuid

@celery_app.task
def mail_daily_inventory_summary():

    product_records = db.query(InventoryItem).all()

    if not product_records:
        return "No inventory items found."

    wb = Workbook()
    current_ws = wb.active
    current_ws.title = "Inventory Summary"

    # Header
    current_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    current_ws.cell(row=1, column=1).value = "All Warehouses - Inventory Summary"
    current_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    current_ws.append(["Product ID", "SKU", "Warehouse ID", "Bin ID", "Qty"])

    for item in product_records:
        current_ws.append([
            str(item.product_id),
            item.sku,
            str(item.warehouse_id),
            str(item.bin_id),
            item.qty
        ])


    reports_dir = "daily_reports"
    os.makedirs(reports_dir, exist_ok=True)
    file_name = f"inventory_summary_{uuid.uuid4()}.xlsx"
    file_path = os.path.join(reports_dir, file_name)
    wb.save(file_path)


    email_obj = Email()
    email_obj.send_email("Daily Inventory Summary", "Daily Inventory Summary","ys6244864@gmail.com",file_path)


    # now = datetime.now().strftime("%H:%M:%S")
    # print(f"Weekly mail task running! Time: {now}")

